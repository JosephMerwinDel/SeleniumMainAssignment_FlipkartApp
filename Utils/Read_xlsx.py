import openpyxl

class XlsxReader:
    @staticmethod
    def get_address_data():
        wb = openpyxl.load_workbook(f"/Users/jmerwin/PycharmProjects/SeleniumMainProject/Utils/AddressDetails.xlsx")
        sheet = wb.active
        row = sheet.max_row
        column = sheet.max_column
        data = {}
        i=2
        data["name"] = sheet.cell(row=i, column=1).value
        data["mobile_no"] = sheet.cell(row=i, column=2).value
        data["pincode"] = sheet.cell(row=i, column=3).value
        data["locality"] = sheet.cell(row=i, column=4).value
        data["address"] = sheet.cell(row=i, column=5).value
        data["address_type"] = sheet.cell(row=i, column=6).value
        return data

